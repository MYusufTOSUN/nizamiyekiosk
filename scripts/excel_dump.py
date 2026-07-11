# SATINALMA_LISTESI_v3.xlsx icerigini tam doker (tum sayfalar, tum satirlar).
from pathlib import Path

import openpyxl

wb = openpyxl.load_workbook(Path(__file__).parent.parent / "SATINALMA_LISTESI_v3.xlsx", data_only=True)
out = []
for ws in wb.worksheets:
    out.append(f"===== SAYFA: {ws.title}  ({ws.max_row} satir x {ws.max_column} sutun) =====")
    for row in ws.iter_rows(values_only=True):
        if all(c is None for c in row):
            continue
        cells = ["" if c is None else str(c).replace("\n", " / ") for c in row]
        out.append(" | ".join(cells))
Path(__file__).parent.parent / "excel_dump.txt"
dump = Path(__file__).parent.parent / "excel_dump.txt"
dump.write_text("\n".join(out), encoding="utf-8")
print(f"yazildi: {dump} ({len(out)} satir)")
