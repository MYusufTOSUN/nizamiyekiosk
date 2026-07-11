# v4 kozmetik yama: cerceve kalem ADI (300x300/18kg) + bayat mukerrer toplam satiri.
# Kullanim: Excel'de dosya KAPALIYKEN  .venv\Scripts\python.exe scripts\excel_v4_patch.py
from pathlib import Path

import openpyxl

p = Path(__file__).parent.parent / "SATINALMA_LISTESI_v4.xlsx"
wb = openpyxl.load_workbook(p)
ws = wb.worksheets[0]
for r in range(2, ws.max_row + 1):
    ad = ws.cell(r, 3).value
    if ad and "VESA 400x400" in str(ad):
        ws.cell(r, 3).value = (
            "VESA 300x300 (oneri: universal 300+400 delikli) yatik (yuzu yukari) "
            "celik tasiyici cerceve — 65 inc ~18 kg (TCL 65C7K), tam kenar destekli"
        )
    if ad and "~ 60-125 bin TL" in str(ad):
        ws.delete_rows(r, 1)
        break
wb.save(p)
print("yama uygulandi")
