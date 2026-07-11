# Yeni eklenen satirlari (No 29-34) orijinal satirlarla ayni duzene getirir:
# font, kenarlik, hizalama (wrap), satir yuksekligi — referans: mevcut kalem satiri.
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, PatternFill

p = Path(__file__).parent.parent / "SATINALMA_LISTESI_v4.xlsx"
wb = openpyxl.load_workbook(p)
ws = wb.worksheets[0]
NO = 1

def row_of(no: int) -> int:
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, NO).value
        try:
            if v is not None and int(float(str(v).strip())) == no:
                return r
        except (TypeError, ValueError):
            continue
    raise KeyError(no)

ref = row_of(3)  # dolu, tipik bir kalem satiri
ref_h = ws.row_dimensions[ref].height
print("referans satir yuksekligi:", ref_h)

yeni_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
for no in range(29, 35):
    r = row_of(no)
    for c in range(1, 11):
        src = ws.cell(ref, c)
        dst = ws.cell(r, c)
        dst.font = copy(src.font)
        dst.border = copy(src.border)
        dst.number_format = src.number_format
        a = copy(src.alignment)
        dst.alignment = Alignment(
            horizontal=a.horizontal, vertical=a.vertical or "top",
            wrap_text=True, shrink_to_fit=False,
        )
        dst.fill = yeni_fill  # 'YENI' ayirt edici acik turuncu kalsin
    # satir yuksekligi: referansla ayni (None ise otomatik birak)
    if ref_h:
        ws.row_dimensions[r].height = ref_h
    else:
        ws.row_dimensions[r].height = None

# Toplam/alt satirlarda da tasma olmasin
for r in range(2, ws.max_row + 1):
    for c in range(1, 11):
        cell = ws.cell(r, c)
        if cell.value and not cell.alignment.wrap_text:
            a = cell.alignment
            cell.alignment = Alignment(horizontal=a.horizontal, vertical=a.vertical or "top", wrap_text=True)

wb.save(p)
print("format uygulandi")
