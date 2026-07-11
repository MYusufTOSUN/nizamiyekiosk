# v4 duzeltme: Faz-2 kalemleri SATIN ALINACAK statuye geri doner (kullanici karari).
# Hicbir kalem cikarilmaz; analiz sadece bilgi notu olarak kalir.
from pathlib import Path

import openpyxl

p = Path(__file__).parent.parent / "SATINALMA_LISTESI_v4.xlsx"
wb = openpyxl.load_workbook(p)
ws = wb.worksheets[0]
NO, AD, TOPLAM, DURUM, NOT_ = 1, 3, 6, 9, 10

def row_of(no: int) -> int:
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, NO).value
        try:
            if v is not None and int(float(str(v).strip())) == no:
                return r
        except (TypeError, ValueError):
            continue
    raise KeyError(no)

FAZ2 = "SATIN ALINACAK — festivalde kullanilmaz (hazir video), FAZ 2 kalici sergi (canli sohbet) icin simdi aliniyor. "
updates = {
    11: ("STOKTA", FAZ2 + "Stoklu EN UCUZ (%30 ind.). 'Kritik Stok' — bekletmeyin. Dogru varyant 'SE/S' (kablo yandan cikisli, panel-montaj); '-X' preamp'siz varyanti ALMAYIN."),
    13: ("STOKTA", FAZ2 + "DOGRULANDI: 821,59 TL, stokta (1-3 gun). 1 asil + 1 YEDEK (2 adet). Muadil: Kirlin MPC-470 10m ~800-860 TL, Kozmos KCL-020-10M ~840 TL."),
    15: ("STOKTA", FAZ2 + "4 adet. DOGRULANDI: 21,05 TL, stokta ('Kritik Stok'). Shure mikrofonla ayni sipariste alin."),
    16: ("OPSIYONEL", FAZ2 + "Yedek mikrofon (OPSIYONEL, butce kisilirsa cikarilabilir). DIKKAT: '-X' varyanti (CVG18-B/C-X) preamp'SIZ, birebir muadil DEGIL — ALMAYIN."),
    19: ("STOKTA", "SATIN ALINACAK. BILGI NOTU: 'Kurallar' sayfasi 'PAR menuden sabit amber (kontrolcusuz)' da diyor — kontrolcu, yavas renk gecisli OTOMATIK ambiyans icin (standalone, guc gelince kendi baslar; 28 saat gozetimsiz icin ideal). Ucretsiz Nicolaudie ESA2 ile PC'de bir kez programlanir. elektroniksatis.com 16.649,11 TL; bot korumasi nedeniyle siparis oncesi TELEFONLA stok teyidi. Alternatif (ayni satici, stok teyitli): SSP PILOT 2000 konsol 22.650 TL artises — ama manuel, guc kesilince otomatik baslamaz."),
    20: ("STOKTA", "SATIN ALINACAK (kontrolcuyle birlikte). Kontrolcu -> PAR1 -> PAR2 -> PAR3 -> PAR4 zinciri: 4 baglanti + 1 YEDEK = 5 adet. Kisa (~1,5-3m) 3-pin XLR DMX. NOT: kontrolcu kabin DISINDAYSA ilk PAR'a ~5-10m 1 kablo gerekebilir. Profesyonel secenek: Procab CAB901 (Meteor Muzik)."),
}
for no, (durum, note) in updates.items():
    r = row_of(no)
    ws.cell(r, DURUM).value = durum
    ws.cell(r, NOT_).value = note

# Toplami yeniden hesapla (kesin fiyatli: STOKTA/KOSULLU/SERBEST)
total_row = None
for r in range(2, ws.max_row + 1):
    v = ws.cell(r, AD).value
    if v and "ÇEKİRDEK TL TOPLAMI" in str(v):
        total_row = r
        break
core = 0.0
for r in range(2, total_row):
    durum = str(ws.cell(r, DURUM).value or "")
    top = ws.cell(r, TOPLAM).value
    if isinstance(top, (int, float)) and durum in ("STOKTA", "KOSULLU", "SERBEST"):
        core += float(top)
ws.cell(total_row, TOPLAM).value = round(core, 2)
ws.cell(total_row, AD).value = "ÇEKİRDEK TL TOPLAMI v4 (kesin fiyatlı; teklif/opsiyonel/teyit-bekleyen hariç — HİÇBİR KALEM ÇIKARILMADI)"
ws.cell(total_row, NOT_).value = (
    "v3: 699.463,57 -> v4: TV degisimi -30.399 (C7K); YENI dokunmatik/guvenlik kalemleri +9.750 "
    "(USB uzatma, 3. HDMI, bariyer, yedek bellek; dokunmatik monitor fiyat TEYIT bekliyor, haric). "
    "Mikrofon zinciri ve DMX seti LISTEDE ve toplamda (kullanici karari: Faz 2 dahil tek alim)."
)
wb.save(p)
print(f"geri alindi; cekirdek v4: {core:,.2f} TL")
